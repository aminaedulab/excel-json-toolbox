import streamlit as st
import pandas as pd
import zipfile
from io import BytesIO
from openpyxl.styles import PatternFill

# --- Split Excel ---
def split_excel_ui():
    st.subheader("📂 Split Excel File")

    file = st.file_uploader("Upload Excel File", type=['xlsx'])
    entries_per_file = st.number_input("Entries per file", min_value=1, value=100)

    if st.button("Split File") and file:
        df = pd.read_excel(file)
        chunks = [df[i:i + entries_per_file] for i in range(0, len(df), entries_per_file)]
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            for idx, chunk in enumerate(chunks, 1):
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    chunk.to_excel(writer, index=False)
                zipf.writestr(f"part_{idx}.xlsx", buffer.getvalue())

        st.download_button(
            "⬇️ Download ZIP (Split Files)",
            zip_buffer.getvalue(),
            file_name="split_files.zip",
            mime="application/zip"
        )
        st.success(f"✅ Split into {len(chunks)} files.")


# --- Find Duplicates ---
def find_duplicates_ui():
    st.subheader("🔍 Find Duplicate Rows")

    file = st.file_uploader("Upload Excel File", type=['xlsx'])
    if file:
        df = pd.read_excel(file)
        cols = st.multiselect("Select columns to check for duplicates", df.columns)
        if st.button("Find Duplicates") and cols:
            dups = df[df.duplicated(subset=cols, keep=False)]
            st.write(f"Found {len(dups)} duplicate rows.")

            yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
                sheet = writer.sheets["Sheet1"]
                for col in cols:
                    col_idx = df.columns.get_loc(col) + 1
                    dup_vals = df[df.duplicated(subset=[col], keep=False)][col].unique()
                    for r_idx, val in enumerate(df[col], start=2):
                        if val in dup_vals:
                            sheet.cell(row=r_idx, column=col_idx).fill = yellow

            st.download_button(
                "⬇️ Download Highlighted Excel",
                output.getvalue(),
                file_name="highlighted_duplicates.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


# --- Merge Excel ---
def merge_excel_ui():
    st.subheader("🧩 Merge Excel Files or Sheets")

    merge_type = st.radio(
        "Choose merge type:",
        ["📁 Merge Multiple Excel Files", "📑 Merge All Sheets in One File"]
    )

    if merge_type == "📁 Merge Multiple Excel Files":
        files = st.file_uploader("Upload Multiple Excel Files", type=['xlsx'], accept_multiple_files=True)
        if files and st.button("Merge Files"):
            merged_data = []
            for file in files:
                df = pd.read_excel(file)
                df["Source_File"] = file.name  # add file name for traceability
                merged_data.append(df)

            merged_df = pd.concat(merged_data, ignore_index=True)
            st.dataframe(merged_df.head())

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                merged_df.to_excel(writer, index=False, sheet_name="Merged_Data")

            st.download_button(
                "⬇️ Download Merged Excel File",
                buffer.getvalue(),
                file_name="merged_files.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"✅ Merged {len(files)} Excel files successfully!")

    elif merge_type == "📑 Merge All Sheets in One File":
        file = st.file_uploader("Upload Excel File (with multiple sheets)", type=['xlsx'])
        if file and st.button("Merge Sheets"):
            sheet_dict = pd.read_excel(file, sheet_name=None)
            merged_df = pd.concat(sheet_dict.values(), ignore_index=True)
            st.dataframe(merged_df.head())

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                merged_df.to_excel(writer, index=False, sheet_name="Merged_Sheets")

            st.download_button(
                "⬇️ Download Merged Excel File",
                buffer.getvalue(),
                file_name="merged_sheets.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"✅ Merged {len(sheet_dict)} sheets successfully!")
